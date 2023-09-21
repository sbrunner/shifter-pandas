"""Datasource builder for data from British Petroleum."""

from typing import Any, Optional

import openpyxl
import pandas as pd

from shifter_pandas import standardize_property
from shifter_pandas.wikidata_ import WikidataDatasource

# ISO
UNITS_ENERGY = ["J", "J (input-equivalent)"]
# ISO
UNITS_ENERGY_PER_CAPITA = ["J per capita"]
# ISO
UNITS_VOLUME = ["m³"]
# ISO
UNITS_VOLUME_PRE_DAY = ["m³ per day"]
# ISO
UNITS_MASS = ["tonnes"]


class BPDatasource:
    """Datasource builder for data from British Petroleum."""

    def __init__(self, file_name: str) -> None:
        """Initialize the datasource builder."""
        self.xlsx = openpyxl.load_workbook(file_name)
        self.wdds = WikidataDatasource()
        self.wdds.set_alias("World", "World", "Q16502", "World")
        # Crude oil: oil_units_conversion[<from unit>][<to unit>] = <factor>
        self.oil_units_conversion: dict[str, dict[str, float]] = {}
        # Oil products: oil_products_units_conversion[<product>][<from unit>][<to unit>] = <factor>
        self.oil_products_units_conversion: dict[str, dict[str, dict[str, float]]] = {}
        # Natural gas (NG) and liquefied natural gas (LNG):
        # gaz_units_conversion[<from unit>][<to unit>] = <factor>
        self.gaz_units_conversion: dict[str, dict[str, float]] = {}

        # to_iso_unit[<from_unit>] = {"unit": <to_unit>, "factor": <factor>}
        # tones is considered as ISO because it's more convenient to use
        # then J, m³, tonnes (day, capita).
        self.to_iso_unit: dict[str, dict[str, Any]] = {}

        # 1 metric tonne = 2204.62 lb.
        self.to_iso_unit["lb"] = {"unit": "tonnes", "factor": 2204.62}
        # = 1.1023 short tons
        self.to_iso_unit["short tons"] = {"unit": "tonnes", "factor": 1.1023}
        # 1 kilolitre = 6.2898 barrels
        self.to_iso_unit["barrels"] = {"unit": "m³", "factor": 6.2898}
        # 1 kilolitre = 1 cubic meter
        self.to_iso_unit["litres"] = {"unit": "m³", "factor": 1000}
        # 1 kilocalorie (kcal) = 4.1868 kJ = 3.968 Btu
        # 1 kilojoule (kJ) = 1,000 joules = 0.239 kcal = 0.948 Btu
        self.to_iso_unit["kilocalorie"] = {"unit": "J", "factor": 1 / 239}
        self.to_iso_unit["kcal"] = self.to_iso_unit["kilocalorie"]
        self.to_iso_unit["calorie"] = {"unit": "J", "factor": 1 / 0.239}
        self.to_iso_unit["cal"] = self.to_iso_unit["calorie"]
        self.to_iso_unit["Btu"] = {"unit": "J", "factor": 1 / 948}
        # 1 British thermal unit (Btu) = 0.252 kcal = 1.055 kJ
        # 1 barrel of oil equivalent (boe) = 5.8 million Btu = 6.119 million kJ
        self.to_iso_unit["barrel of oil equivalent"] = {"unit": "j", "factor": 6119000}
        # 1 kilowatt-hour (kWh) = 860 kcal = 3600 kJ = 3412 Btu
        self.to_iso_unit["kilowatt-hour"] = {"unit": "J", "factor": 3600 / 1000}
        self.to_iso_unit["kilowatt-hours"] = self.to_iso_unit["kilowatt-hour"]
        self.to_iso_unit["kWh"] = self.to_iso_unit["kilowatt-hour"]
        self.to_iso_unit["watt-hour"] = {"unit": "J", "factor": 3600}
        self.to_iso_unit["watt-hours"] = self.to_iso_unit["watt-hour"]
        self.to_iso_unit["Wh"] = self.to_iso_unit["watt-hour"]

        self.to_iso_unit["cubic meters"] = {"unit": "m³", "factor": 1}
        self.to_iso_unit["meters"] = {"unit": "m", "factor": 1}
        self.to_iso_unit["joules"] = {"unit": "J", "factor": 1}
        self.to_iso_unit["cubic feets"] = {"unit": "m³", "factor": 35.3146667}
        self.to_iso_unit["cubic meter"] = self.to_iso_unit["cubic meters"]
        self.to_iso_unit["meter"] = self.to_iso_unit["meters"]
        self.to_iso_unit["joule"] = self.to_iso_unit["joules"]
        self.to_iso_unit["cubic feet"] = self.to_iso_unit["cubic feets"]
        self.to_iso_unit["watts"] = {"unit": "W", "factor": 1}
        self.to_iso_unit["watt"] = self.to_iso_unit["watts"]

        for type_index, type_value in enumerate(self.xlsx.sheetnames):
            if type_value == "Approximate conversion factors":
                self.units_sheet = self.xlsx.worksheets[type_index]

        for raw in range(8, 13):
            from_unit = self.normalize_unit(self.units_sheet.cell(raw, 1).value)
            from_iso_unit, _, from_iso_factor = self._iso_unit(from_unit)
            for col in range(4, 9):
                to_unit_1 = self.units_sheet.cell(4, col).value
                to_unit = self.normalize_unit(
                    ((to_unit_1 + " ") if to_unit_1 is not None else "") + self.units_sheet.cell(5, col).value
                )
                to_iso_unit, _, to_iso_factor = self._iso_unit(to_unit)
                value = self.units_sheet.cell(raw, col).value
                if isinstance(value, (int, float)) and from_unit != to_unit:
                    self.oil_units_conversion.setdefault(from_unit, {})[to_unit] = value
                    self.oil_units_conversion.setdefault(from_iso_unit, {})[to_iso_unit] = (
                        value / from_iso_factor * to_iso_factor
                    )

        self.to_iso_unit["us gallons"] = {
            "unit": "m³",
            "factor": self.oil_units_conversion["us gallons"]["m³"],
        }
        self.to_iso_unit["gallons"] = self.to_iso_unit["us gallons"]

        for raw in range(20, 27):
            product = self.units_sheet.cell(raw, 1).value
            for col in range(3, 9):
                from_unit = self.normalize_unit(self.units_sheet.cell(16, col).value)
                from_iso_unit, _, from_iso_factor = self._iso_unit(from_unit)
                to_unit = self.normalize_unit(self.units_sheet.cell(17, col).value[2:])
                to_iso_unit, _, to_iso_factor = self._iso_unit(to_unit)
                value = self.units_sheet.cell(raw, col).value
                if isinstance(value, (int, float)) and from_unit != to_unit:
                    self.oil_products_units_conversion.setdefault(product, {}).setdefault(from_unit, {})[
                        to_unit
                    ] = value
                    self.oil_products_units_conversion.setdefault(product, {}).setdefault(from_iso_unit, {})[
                        to_iso_unit
                    ] = (value / from_iso_factor * to_iso_factor)

        for raw in range(33, 39):
            from_unit = self.normalize_unit(self.units_sheet.cell(raw, 1).value)
            from_iso_unit, _, from_iso_factor = self._iso_unit(from_unit)
            for col in range(3, 10):
                to_unit = self.normalize_unit(
                    self.units_sheet.cell(29, col).value + " " + self.units_sheet.cell(30, col).value
                )
                to_iso_unit, _, to_iso_factor = self._iso_unit(to_unit)
                value = self.units_sheet.cell(raw, col).value
                if isinstance(value, (int, float)) and from_unit != to_unit:
                    self.gaz_units_conversion.setdefault(from_unit, {})[to_unit] = value
                    self.gaz_units_conversion.setdefault(from_iso_unit, {})[to_iso_unit] = (
                        value / from_iso_factor * to_iso_factor
                    )

    @staticmethod
    def normalize_unit(unit: str) -> str:
        """Get normalized unit."""
        unit = unit.strip()
        unit = unit.lower()
        if unit.startswith("1 "):
            unit = unit[2:]
        unit = unit.replace("/", " / ")
        unit = unit.replace("  ", " ")
        unit = unit.replace("equiv.", "equivalent")
        unit = unit.replace(" daily", " per day")
        return unit

    def _iso_unit(self, unit: str) -> tuple[str, str, float]:
        if "/" not in unit:
            return self._single_iso_unit(unit)
        else:
            upper, lower = unit.split("/")
            unit, upper_postfix, factor = self._single_iso_unit(upper.strip())
            assert not upper_postfix, f"Upper unit {upper} has a postfix {upper_postfix}"
            lower_unit, postfix, lower_factor = self._single_iso_unit(lower.strip())
            return f"{unit} / {lower_unit}", postfix, factor / lower_factor

    def _single_iso_unit(self, unit: str) -> tuple[str, str, float]:
        unit_postfix = ""
        postfix = ""
        for postfix_candidate in ("*",):
            if unit.endswith(postfix_candidate):
                postfix = postfix_candidate
                unit = unit[: -len(postfix_candidate)]
                break

        for postfix_candidate in (
            " (input-equivalent)",
            " per capita",
            " daily",
            " per year",
            " per day",
        ):
            if unit.endswith(postfix_candidate):
                unit_postfix = postfix_candidate
                unit = unit[: -len(postfix_candidate)]
                break
        for postfix_candidate in (
            " of carbon dioxide",
            " of oil equivalent",
            " of lithium content",
            "1",
            "*",
        ):
            if unit.endswith(postfix_candidate):
                postfix = postfix_candidate
                unit = unit[: -len(postfix_candidate)]
                break

        factor = 1
        if unit.startswith("kilo"):
            unit = unit[4:]
            factor = 1000
        elif unit.startswith("mega"):
            unit = unit[4:]
            factor = 1000000
        elif unit.startswith("giga"):
            unit = unit[4:]
            factor = 1000000000
        elif unit.startswith("tera"):
            unit = unit[4:]
            factor = 1000000000000
        elif unit.startswith("peta"):
            unit = unit[4:]
            factor = 1000000000000000
        elif unit.startswith("exa"):
            unit = unit[3:]
            factor = 1000000000000000000
        elif unit.startswith("thousand million "):
            unit = unit[17:]
            factor = 1000000000
        elif unit.startswith("thousand "):
            unit = unit[9:]
            factor = 1000
        elif unit.startswith("million "):
            unit = unit[8:]
            factor = 1000000
        elif unit.startswith("billion "):
            unit = unit[8:]
            factor = 1000000000
        elif unit.startswith("trillion "):
            unit = unit[9:]
            factor = 1000000000000

        if unit in self.to_iso_unit:
            definition = self.to_iso_unit[unit]
            unit = definition["unit"]
            factor *= definition["factor"]

        return unit + unit_postfix, postfix, factor

    def metadata(self) -> list[dict[str, Any]]:
        """Get the metadata."""
        metadata: list[dict[str, Any]] = []
        for type_index, type_value in enumerate(self.xlsx.sheetnames):
            nice_type = type_value
            for postfix in (
                " - TWh",
                " - EJ",
                " - PJ",
                " - Cons capita",
                " - Barrels",
                " - Tonnes",
                " - Kboed",
                " - Prices",
            ):
                if type_value.endswith(postfix):
                    nice_type = type_value[: -len(postfix)]

            row_index = -1
            for index in (3, 4):
                value = self.xlsx.worksheets[type_index].cell(index, 2).value
                if isinstance(value, int) and 1800 < value < 2100:
                    row_index = index
                    break

            if row_index >= 0:
                # Year
                years = []
                index = 2
                while True:
                    value = self.xlsx.worksheets[type_index].cell(row_index, index).value
                    if self.xlsx.worksheets[type_index].cell(row_index - 1, index).value is not None:
                        break
                    years.append({"label": value, "index": index})
                    index += 1

                # Country
                regions = []
                index = row_index + 2
                nb_empty_cells = 0
                while True:
                    value = self.xlsx.worksheets[type_index].cell(index, 1).value
                    if value is not None:
                        nb_empty_cells = 0
                        regions.append({"label": value, "index": index})
                    nb_empty_cells += 1
                    if nb_empty_cells > 5:
                        break
                    index += 1
                unit = {"original": self.xlsx.worksheets[type_index].cell(row_index, 1).value.strip()}
                unit["normalized"] = self.normalize_unit(unit["original"])
                iso_unit, postfix, factor = self._iso_unit(unit["normalized"])
                unit["iso"] = iso_unit
                unit["iso_factor"] = factor
                unit["iso_postfix"] = postfix
                metadata.append(
                    {
                        "type": type_value,
                        "label": nice_type,
                        "index": type_index,
                        "unit": unit,
                        "years": years,
                        "regions": regions,
                        "supported": True,
                        "row_index": row_index,
                    }
                )
            else:
                metadata.append(
                    {
                        "type": type_value,
                        "index": type_index,
                        "supported": False,
                    }
                )

        return metadata

    def datasource(
        self,
        types_filter: Optional[list[str]] = None,
        regions_filter: Optional[list[str]] = None,
        units_filter: Optional[list[str]] = None,
        years_filter: Optional[list[int]] = None,
        years_factor: Optional[int] = None,
        units: str = "iso",
        wikidata_id: bool = False,
        wikidata_type: bool = False,
        wikidata_name: bool = False,
        wikidata_properties: Optional[list[str]] = None,
    ) -> pd.DataFrame:
        """Get the Datasource as DataFrame."""

        if wikidata_properties is None:
            wikidata_properties = []
        wikidata = wikidata_id or wikidata_name or wikidata_properties

        columns = ["Value", "Type", "Unit", "TypeUnit", "Year", "Region"]
        if wikidata:
            if wikidata_id:
                columns.append("WikidataId")
            if wikidata_name:
                columns.append("WikidataName")
            if wikidata_type:
                columns.append("WikidataType")
            for wikidata_property in wikidata_properties:
                columns.append(
                    f"Wikidata{standardize_property(self.wdds.get_property_name(wikidata_property))}"
                )
        data_frame = pd.DataFrame(columns=columns)
        for type_ in self.metadata():
            if not type_["supported"]:
                continue
            type_index = type_["index"]
            type_type = type_["type"]
            type_label = type_["label"]
            years = type_["years"]
            regions = type_["regions"]

            if types_filter is not None and type_type not in types_filter:
                continue

            for year in years:
                if years_filter is not None and year["label"] not in years_filter:
                    continue
                if years_factor is not None and year["label"] % years_factor != 0:
                    continue
                for region in regions:
                    if regions_filter is not None and region["label"] not in regions_filter:
                        continue
                    value = self.xlsx.worksheets[type_index].cell(region["index"], year["index"]).value
                    if not isinstance(value, int) and not isinstance(value, float):
                        continue

                    unit_definition = type_["unit"]
                    unit_postfix = ""
                    if units == "normalized":
                        unit = unit_definition["normalized"]
                    elif units == "iso":
                        unit = unit_definition["iso"]
                        value = value * unit_definition["iso_factor"]
                        unit_postfix = unit_definition["iso_postfix"]
                    else:
                        unit = unit_definition["original"]

                    if units_filter is not None and unit not in units_filter:
                        continue

                    element = {
                        "Value": value,
                        "Year": year["label"],
                        "Region": region["label"],
                        "Type": type_type,
                        "Unit": f"{unit}{unit_postfix}",
                        "TypeUnit": f"{type_label} [{unit}]{unit_postfix}",
                    }
                    if wikidata:
                        region_label = region["label"]
                        if region_label.startswith("Total "):
                            region_label = region_label[6:]
                        element_id = self.wdds.get_region(region_label)
                        element["WikidataType"] = element_id["type"] if element_id else None
                        element.update(
                            self.wdds.get_item(
                                element_id["id"] if element_id else None,
                                with_name=wikidata_name,
                                with_id=wikidata_id,
                                properties=wikidata_properties,
                                prefix="Wikidata",
                            )
                        )
                    data_frame = pd.concat(
                        [data_frame, pd.DataFrame({k: [v] for k, v in element.items()})], ignore_index=True
                    )

        return data_frame

    def datasource_non_fossil_electricity_to_primary_energy_factor(
        self, from_year: int = 1900
    ) -> pd.DataFrame:
        """Get the Datasource used to convert non fossil electricity to primary energy as DataFrame."""

        data: dict[str, list[float]] = {
            "Year": [],
            "Factor": [],
        }
        before_2001 = self.units_sheet["B45"].value
        for year in range(from_year, 2001):
            data["Year"].append(year)
            data["Factor"].append(before_2001)
        for index, year in enumerate(range(2001, 2010)):
            data["Year"].append(year)
            data["Factor"].append(self.units_sheet.cell(index + 46, 2).value)
        for index, year in enumerate(range(2010, 2020)):
            data["Year"].append(year)
            data["Factor"].append(self.units_sheet.cell(index + 45, 5).value)

        return pd.DataFrame(data)
